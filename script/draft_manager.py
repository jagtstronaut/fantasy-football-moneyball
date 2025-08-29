#!/usr/bin/env python3

import pandas as pd
import openpyxl
import os
import sys
from typing import Optional

class FantasyDraftManager:
    def __init__(self, spreadsheet_path: str):
        self.spreadsheet_path = spreadsheet_path
        self.sheets_data = {}
        self.pending_deletions = {}  # Track rows to delete: {sheet_name: [row_indices]}
        self.my_squad_updates = {}  # Track position updates: {position: count_change}
        self.load_spreadsheet()
    
    def load_spreadsheet(self):
        """Load all sheets from the Excel file"""
        try:
            # First, restore from backup if it exists
            self._restore_from_backup()
            
            print(f"Loading spreadsheet: {self.spreadsheet_path}")
            # Read all sheets
            excel_file = pd.ExcelFile(self.spreadsheet_path)
            print(f"Found sheets: {excel_file.sheet_names}")
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(self.spreadsheet_path, sheet_name=sheet_name)
                self.sheets_data[sheet_name] = df
                print(f"Loaded {sheet_name}: {len(df)} rows, {len(df.columns)} columns")
                
        except Exception as e:
            print(f"Error loading spreadsheet: {e}")
            sys.exit(1)
    
    def _restore_from_backup(self):
        """Restore the main spreadsheet from backup at startup"""
        backup_path = self.spreadsheet_path.replace('.xlsx', '_backup.xlsx')
        
        if os.path.exists(backup_path):
            try:
                import shutil
                shutil.copy2(backup_path, self.spreadsheet_path)
                print(f"✅ Restored from backup: {os.path.basename(backup_path)} → {os.path.basename(self.spreadsheet_path)}")
            except Exception as e:
                print(f"Warning: Could not restore from backup: {e}")
        else:
            print(f"ℹ️  No backup file found at {backup_path} - using existing spreadsheet")
    
    def display_available_players(self, sheet_name: str, max_rows: int = 5):
        """Display available players from a specific sheet"""
        if sheet_name not in self.sheets_data:
            print(f"Sheet '{sheet_name}' not found!")
            return
        
        df = self.sheets_data[sheet_name]
        print(f"\n=== {sheet_name.upper()} ===")
        print(f"Total players: {len(df)}")
        
        if len(df) == 0:
            print("No players remaining in this sheet!")
            return
        
        # Show column names
        print("\nColumns:", ", ".join([str(col) for col in df.columns.tolist()]))
        
        # Display first few rows
        print(f"\nFirst {min(max_rows, len(df))} players:")
        print(df.head(max_rows).to_string(index=False))
        
        if len(df) > max_rows:
            print(f"... and {len(df) - max_rows} more players")
    
    def search_player_by_last_name(self, last_name: str) -> list:
        """Search for players by last name across all sheets"""
        matches = []
        last_name_lower = last_name.lower().strip()
        
        for sheet_name, df in self.sheets_data.items():
            if len(df) == 0:
                continue
                
            # Look through all columns for potential player names
            for col in df.columns:
                if df[col].dtype == 'object':  # Text columns
                    # Search for last name in this column
                    mask = df[col].astype(str).str.lower().str.contains(last_name_lower, na=False, regex=False)
                    matching_rows = df[mask]
                    
                    for idx, row in matching_rows.iterrows():
                        matches.append({
                            'sheet': sheet_name,
                            'index': idx,
                            'row_data': row,
                            'matching_column': col
                        })
        
        return matches
    
    def remove_player_rows(self, matches: list, indices_to_remove: list):
        """Track player rows for removal (actual deletion happens during save)"""
        removed_count = 0
        
        for i in indices_to_remove:
            if 0 <= i < len(matches):
                match = matches[i]
                sheet_name = match['sheet']
                row_index = match['index']
                
                # Track this row for deletion
                if sheet_name not in self.pending_deletions:
                    self.pending_deletions[sheet_name] = []
                
                # Convert pandas index to Excel row (add 2 because Excel is 1-indexed and has header)
                excel_row = row_index + 2
                self.pending_deletions[sheet_name].append(excel_row)
                
                removed_count += 1
                print(f"Marked player from {sheet_name} for removal")
        
        return removed_count
    
    def pick_player_for_team(self, matches: list, indices_to_pick: list):
        """Pick players for your team (removes from available and updates squad count)"""
        picked_count = 0
        
        for i in indices_to_pick:
            if 0 <= i < len(matches):
                match = matches[i]
                sheet_name = match['sheet']
                row_index = match['index']
                
                # Track this row for deletion (same as remove_player_rows)
                if sheet_name not in self.pending_deletions:
                    self.pending_deletions[sheet_name] = []
                
                excel_row = row_index + 2
                self.pending_deletions[sheet_name].append(excel_row)
                
                # Update squad count based on sheet
                position_map = {
                    'QBs': 'QB',
                    'RBs': 'RB', 
                    'WRs': 'WR',
                    'Ks': 'K',
                    'Ds': 'D',
                    'TEs': 'TE'
                }
                
                if sheet_name in position_map:
                    position = position_map[sheet_name]
                    if position not in self.my_squad_updates:
                        self.my_squad_updates[position] = 0
                    self.my_squad_updates[position] += 1
                
                picked_count += 1
                print(f"Picked player from {sheet_name} for your team!")
        
        return picked_count
    
    def update_slip_values(self):
        """Update slip values for positions in Decision Matrix"""
        if 'Decision Matrix' not in self.sheets_data:
            print("Decision Matrix not found!")
            return
        
        print("\n=== UPDATE SLIP VALUES ===")
        print("Enter new slip values for each position (or press Enter to skip):")
        
        positions = ['QB', 'RB', 'WR', 'K', 'D', 'TE']
        slip_updates = {}
        
        for pos in positions:
            try:
                current_slip = input(f"{pos} slip value: ").strip()
                if current_slip:
                    slip_updates[pos] = int(current_slip)
            except ValueError:
                print(f"Invalid value for {pos}, skipping...")
        
        if not slip_updates:
            print("No updates to make.")
            return
        
        # Track the updates for saving
        if not hasattr(self, 'slip_updates'):
            self.slip_updates = {}
        self.slip_updates.update(slip_updates)
        
        print(f"Slip values updated: {slip_updates}")
        return slip_updates
    
    def save_spreadsheet(self):
        """Save the updated spreadsheet using openpyxl to preserve formulas"""
        has_slip_updates = hasattr(self, 'slip_updates') and self.slip_updates
        if not self.pending_deletions and not self.my_squad_updates and not has_slip_updates:
            print("No changes to save")
            return
            
        try:
            # Load workbook with openpyxl to preserve formulas
            workbook = openpyxl.load_workbook(self.spreadsheet_path)
            
            # Process deletions for each sheet
            for sheet_name, rows_to_delete in self.pending_deletions.items():
                if sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # Sort rows in descending order to delete from bottom up
                    # This prevents row numbers from shifting during deletion
                    rows_to_delete.sort(reverse=True)
                    
                    for row_num in rows_to_delete:
                        worksheet.delete_rows(row_num)
                        print(f"Deleted row {row_num} from {sheet_name}")
            
            # Update Decision Matrix with new squad counts
            if self.my_squad_updates and 'Decision Matrix' in workbook.sheetnames:
                dm_sheet = workbook['Decision Matrix']
                
                # Find the "My squad" row (look for text in column A)
                for row_num in range(1, dm_sheet.max_row + 1):
                    cell_value = dm_sheet.cell(row=row_num, column=1).value
                    if cell_value and str(cell_value).strip().lower() == 'my squad':
                        # Column mapping: B=QB(1), C=RB(2), D=WR(3), E=K(4), F=D(5), G=TE(6)
                        pos_to_col = {'QB': 2, 'RB': 3, 'WR': 4, 'K': 5, 'D': 6, 'TE': 7}
                        
                        for position, count_change in self.my_squad_updates.items():
                            if position in pos_to_col:
                                col_num = pos_to_col[position]
                                current_cell = dm_sheet.cell(row=row_num, column=col_num)
                                current_value = str(current_cell.value) if current_cell.value else "0/0"
                                
                                # Parse current format "X/Y"
                                if '/' in current_value:
                                    current_picked, total_limit = current_value.split('/')
                                    new_picked = int(current_picked) + count_change
                                    new_value = f"{new_picked}/{total_limit}"
                                    current_cell.value = new_value
                                    print(f"Updated {position}: {current_value} → {new_value}")
                        break
            
            # Update slip values
            if hasattr(self, 'slip_updates') and self.slip_updates and 'Decision Matrix' in workbook.sheetnames:
                dm_sheet = workbook['Decision Matrix']
                
                # Find the "Slip" row
                for row_num in range(1, dm_sheet.max_row + 1):
                    cell_value = dm_sheet.cell(row=row_num, column=1).value
                    if cell_value and str(cell_value).strip().lower() == 'slip':
                        pos_to_col = {'QB': 2, 'RB': 3, 'WR': 4, 'K': 5, 'D': 6, 'TE': 7}
                        
                        for position, new_value in self.slip_updates.items():
                            if position in pos_to_col:
                                col_num = pos_to_col[position]
                                current_cell = dm_sheet.cell(row=row_num, column=col_num)
                                old_value = current_cell.value
                                current_cell.value = new_value
                                print(f"Updated {position} slip: {old_value} → {new_value}")
                        break
            
            # Save the workbook
            workbook.save(self.spreadsheet_path)
            workbook.close()
            
            # Update our pandas data to reflect the changes
            self._refresh_pandas_data()
            
            # Clear pending changes
            self.pending_deletions = {}
            self.my_squad_updates = {}
            if hasattr(self, 'slip_updates'):
                self.slip_updates = {}
            
            print(f"Spreadsheet updated: {self.spreadsheet_path}")
            
        except Exception as e:
            print(f"Error saving spreadsheet: {e}")
    
    def _refresh_pandas_data(self):
        """Reload pandas data after openpyxl modifications"""
        try:
            excel_file = pd.ExcelFile(self.spreadsheet_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(self.spreadsheet_path, sheet_name=sheet_name)
                self.sheets_data[sheet_name] = df
        except Exception as e:
            print(f"Warning: Could not refresh pandas data: {e}")
    
    def show_menu(self):
        """Display the main menu options"""
        print("\n" + "="*50)
        print("FANTASY FOOTBALL DRAFT MANAGER")
        print("="*50)
        print("1. Show all available players")
        print("2. Show players from specific sheet")
        print("3. Remove player by last name")
        print("4. Pick player for your team")
        print("5. Update slip values")
        print("6. Show summary")
        print("7. Quit")
        print("-"*50)
    
    def _get_calculated_decision_matrix_data(self):
        """Get Decision Matrix data with calculated formula values using openpyxl"""
        try:
            # Load workbook with data_only=True to get calculated values
            workbook = openpyxl.load_workbook(self.spreadsheet_path, data_only=True)
            dm_sheet = workbook['Decision Matrix']
            
            positions = ['QB', 'RB', 'WR', 'K', 'D', 'TE']
            pos_to_col = {1: 'QB', 2: 'RB', 3: 'WR', 4: 'K', 5: 'D', 6: 'TE'}
            
            data = {
                'my_squad': {},
                'top_player': {},
                'slip': {},
                'lower_player': {},
                'diff': {}
            }
            
            # Read data from specific rows
            for row_num in range(1, dm_sheet.max_row + 1):
                cell_value = dm_sheet.cell(row=row_num, column=1).value
                if cell_value:
                    row_label = str(cell_value).strip().lower()
                    
                    if row_label in data:
                        for col_idx, pos in pos_to_col.items():
                            cell = dm_sheet.cell(row=row_num, column=col_idx + 1)  # +1 because pos_to_col is 1-indexed for Excel columns B=2, etc.
                            data[row_label][pos] = cell.value if cell.value is not None else 'N/A'
            
            workbook.close()
            return data
            
        except Exception as e:
            print(f"Warning: Could not read calculated values from Decision Matrix: {e}")
            return None
    
    def _compute_decision_matrix_values(self):
        """Compute Top Player, Lower Player, and Diff values using Python logic"""
        positions = ['QB', 'RB', 'WR', 'K', 'D', 'TE']
        sheet_map = {'QB': 'QBs', 'RB': 'RBs', 'WR': 'WRs', 'K': 'Ks', 'D': 'Ds', 'TE': 'TEs'}
        
        computed_data = {
            'top_player': {},
            'lower_player': {},
            'diff': {}
        }
        
        for pos in positions:
            sheet_name = sheet_map[pos]
            if sheet_name in self.sheets_data:
                df = self.sheets_data[sheet_name]
                
                if len(df) > 0:
                    # Find the points/value column
                    points_col = None
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if any(term in col_lower for term in ['point', 'projected', 'season']):
                            points_col = col
                            break
                    
                    if points_col is not None:
                        # Sort by points descending
                        df_sorted = df.sort_values(by=points_col, ascending=False, na_position='last')
                        df_sorted = df_sorted[pd.notna(df_sorted[points_col])]
                        
                        if len(df_sorted) > 0:
                            # Top player (highest points)
                            top_points = df_sorted.iloc[0][points_col]
                            computed_data['top_player'][pos] = f"{top_points:.0f}" if pd.notna(top_points) else 'N/A'
                            
                            # Get slip value for this position
                            slip_value = self._get_slip_value(pos)
                            
                            if slip_value is not None and slip_value > 0 and len(df_sorted) > slip_value:
                                # Lower player (slip positions down)
                                lower_points = df_sorted.iloc[slip_value][points_col]
                                computed_data['lower_player'][pos] = f"{lower_points:.0f}" if pd.notna(lower_points) else 'N/A'
                                
                                # Diff (top - lower)
                                if pd.notna(top_points) and pd.notna(lower_points):
                                    diff = top_points - lower_points
                                    computed_data['diff'][pos] = f"{diff:.0f}"
                                else:
                                    computed_data['diff'][pos] = 'N/A'
                            else:
                                computed_data['lower_player'][pos] = 'N/A'
                                computed_data['diff'][pos] = 'N/A'
                        else:
                            computed_data['top_player'][pos] = 'N/A'
                            computed_data['lower_player'][pos] = 'N/A'
                            computed_data['diff'][pos] = 'N/A'
                    else:
                        computed_data['top_player'][pos] = 'N/A'
                        computed_data['lower_player'][pos] = 'N/A'
                        computed_data['diff'][pos] = 'N/A'
                else:
                    computed_data['top_player'][pos] = 'N/A'
                    computed_data['lower_player'][pos] = 'N/A'
                    computed_data['diff'][pos] = 'N/A'
        
        return computed_data
    
    def _get_slip_value(self, position):
        """Get the slip value for a position from Decision Matrix"""
        if 'Decision Matrix' in self.sheets_data:
            dm_df = self.sheets_data['Decision Matrix']
            pos_indices = {1: 'QB', 2: 'RB', 3: 'WR', 4: 'K', 5: 'D', 6: 'TE'}
            
            for i, row in dm_df.iterrows():
                if pd.notna(row.iloc[0]) and str(row.iloc[0]).strip().lower() == 'slip':
                    for col_idx, pos in pos_indices.items():
                        if pos == position:
                            slip_val = row.iloc[col_idx]
                            if pd.notna(slip_val):
                                try:
                                    return int(float(slip_val))
                                except (ValueError, TypeError):
                                    return None
        return None

    def show_summary(self):
        """Show summary with Decision Matrix and remaining players by sheet"""
        print("\n=== DRAFT SUMMARY ===")
        
        # Display Decision Matrix data
        if 'Decision Matrix' in self.sheets_data:
            positions = ['QB', 'RB', 'WR', 'K', 'D', 'TE']
            
            # First try to get calculated values from Excel formulas
            calculated_data = self._get_calculated_decision_matrix_data()
            
            # If that fails or returns N/A values, compute them with Python
            computed_data = None
            if not calculated_data or all(calculated_data['top_player'].get(pos) in [None, 'N/A'] for pos in positions):
                computed_data = self._compute_decision_matrix_values()
            
            # Get basic data (My squad, slip) from pandas
            dm_df = self.sheets_data['Decision Matrix']
            my_squad_data = {}
            slip_data = {}
            pos_indices = {1: 'QB', 2: 'RB', 3: 'WR', 4: 'K', 5: 'D', 6: 'TE'}
            
            for i, row in dm_df.iterrows():
                if pd.notna(row.iloc[0]):
                    row_label = str(row.iloc[0]).strip().lower()
                    
                    if row_label == 'my squad':
                        for col_idx, pos in pos_indices.items():
                            my_squad_data[pos] = row.iloc[col_idx] if pd.notna(row.iloc[col_idx]) else '0/0'
                    
                    elif row_label == 'slip':
                        for col_idx, pos in pos_indices.items():
                            slip_data[pos] = row.iloc[col_idx] if pd.notna(row.iloc[col_idx]) else 'N/A'
            
            print("\n🏈 YOUR DRAFT STATUS:")
            
            # Display the data in a nice format
            for pos in positions:
                print(f"\n{pos}:")
                print(f"  Drafted: {my_squad_data.get(pos, '0/0')}")
                
                # Use calculated data if available, otherwise computed data
                if calculated_data and calculated_data['top_player'].get(pos) not in [None, 'N/A']:
                    print(f"  Top Player: {calculated_data['top_player'].get(pos, 'N/A')}")
                    print(f"  Lower Player: {calculated_data['lower_player'].get(pos, 'N/A')}")
                    print(f"  Diff: {calculated_data['diff'].get(pos, 'N/A')}")
                elif computed_data:
                    print(f"  Top Player: {computed_data['top_player'].get(pos, 'N/A')}")
                    print(f"  Lower Player: {computed_data['lower_player'].get(pos, 'N/A')}")
                    print(f"  Diff: {computed_data['diff'].get(pos, 'N/A')}")
                else:
                    print(f"  Top Player: N/A")
                    print(f"  Lower Player: N/A")
                    print(f"  Diff: N/A")
                
                print(f"  Slip: {slip_data.get(pos, 'N/A')}")
        
        print("\n📋 AVAILABLE PLAYERS:")
        total_players = 0
        for sheet_name, df in self.sheets_data.items():
            if sheet_name != 'Decision Matrix':  # Skip Decision Matrix in player count
                count = len(df)
                total_players += count
                print(f"{sheet_name}: {count} players remaining")
        print(f"TOTAL: {total_players} players remaining")
    
    def run(self):
        """Main program loop"""
        print("Welcome to Fantasy Football Draft Manager!")
        print(f"Managing spreadsheet: {os.path.basename(self.spreadsheet_path)}")
        
        while True:
            self.show_menu()
            
            try:
                choice = input("Enter your choice (1-7): ").strip()
                
                if choice == '1':
                    # Show all players (except Decision Matrix)
                    for sheet_name in self.sheets_data.keys():
                        if sheet_name != 'Decision Matrix':
                            self.display_available_players(sheet_name)
                
                elif choice == '2':
                    # Show specific sheet (except Decision Matrix)
                    player_sheets = [name for name in self.sheets_data.keys() if name != 'Decision Matrix']
                    print("\nAvailable player sheets:")
                    for i, sheet_name in enumerate(player_sheets, 1):
                        print(f"{i}. {sheet_name}")
                    
                    try:
                        sheet_choice = int(input("Enter sheet number: ")) - 1
                        if 0 <= sheet_choice < len(player_sheets):
                            self.display_available_players(player_sheets[sheet_choice])
                        else:
                            print("Invalid sheet number!")
                    except ValueError:
                        print("Please enter a valid number!")
                
                elif choice == '3':
                    # Remove player
                    last_name = input("Enter player's last name to remove: ").strip()
                    if not last_name:
                        print("Please enter a last name!")
                        continue
                    
                    matches = self.search_player_by_last_name(last_name)
                    
                    if not matches:
                        print(f"No players found with last name '{last_name}'")
                        continue
                    
                    print(f"\nFound {len(matches)} potential matches:")
                    for i, match in enumerate(matches):
                        row_data = match['row_data']
                        print(f"{i+1}. Sheet: {match['sheet']}")
                        print(f"   Data: {row_data.to_dict()}")
                        print()
                    
                    # Ask which ones to remove
                    remove_input = input("Enter numbers to remove (e.g., 1,3 or 'all' or 'none'): ").strip().lower()
                    
                    if remove_input == 'none':
                        continue
                    elif remove_input == 'all':
                        indices_to_remove = list(range(len(matches)))
                    else:
                        try:
                            indices_to_remove = [int(x.strip()) - 1 for x in remove_input.split(',') if x.strip()]
                        except ValueError:
                            print("Invalid input format!")
                            continue
                    
                    if indices_to_remove:
                        removed_count = self.remove_player_rows(matches, indices_to_remove)
                        if removed_count > 0:
                            print(f"Removed {removed_count} player(s)")
                            self.save_spreadsheet()
                        else:
                            print("No players were removed")
                
                elif choice == '4':
                    # Pick player for team
                    last_name = input("Enter player's last name to pick for your team: ").strip()
                    if not last_name:
                        print("Please enter a last name!")
                        continue
                    
                    matches = self.search_player_by_last_name(last_name)
                    
                    if not matches:
                        print(f"No players found with last name '{last_name}'")
                        continue
                    
                    print(f"\nFound {len(matches)} potential matches:")
                    for i, match in enumerate(matches):
                        row_data = match['row_data']
                        print(f"{i+1}. Sheet: {match['sheet']}")
                        print(f"   Data: {row_data.to_dict()}")
                        print()
                    
                    # Ask which ones to pick
                    pick_input = input("Enter numbers to pick for your team (e.g., 1,3 or 'all' or 'none'): ").strip().lower()
                    
                    if pick_input == 'none':
                        continue
                    elif pick_input == 'all':
                        indices_to_pick = list(range(len(matches)))
                    else:
                        try:
                            indices_to_pick = [int(x.strip()) - 1 for x in pick_input.split(',') if x.strip()]
                        except ValueError:
                            print("Invalid input format!")
                            continue
                    
                    if indices_to_pick:
                        picked_count = self.pick_player_for_team(matches, indices_to_pick)
                        if picked_count > 0:
                            print(f"Picked {picked_count} player(s) for your team!")
                            self.save_spreadsheet()
                        else:
                            print("No players were picked")
                
                elif choice == '5':
                    # Update slip values
                    slip_updates = self.update_slip_values()
                    if slip_updates:
                        self.save_spreadsheet()
                
                elif choice == '6':
                    # Show summary
                    self.show_summary()
                
                elif choice == '7':
                    # Quit
                    print("Thanks for using Fantasy Football Draft Manager!")
                    break
                
                else:
                    print("Invalid choice! Please enter 1-7.")
            
            except KeyboardInterrupt:
                print("\n\nGoodbye!")
                break
            except Exception as e:
                print(f"An error occurred: {e}")
                print("Please try again.")

def main():
    spreadsheet_path = "/Users/jjagt/hacker/ff-moneyball/FF - main.xlsx"
    
    if not os.path.exists(spreadsheet_path):
        print(f"Error: Spreadsheet not found at {spreadsheet_path}")
        return
    
    manager = FantasyDraftManager(spreadsheet_path)
    manager.run()

if __name__ == "__main__":
    main()